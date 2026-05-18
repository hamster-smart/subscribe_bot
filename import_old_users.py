"""
Скрипт импорта подписчиков из старого бота.

Запуск на сервере:
    docker compose cp import_old_users.py bot:/app/
    docker compose exec bot python3 import_old_users.py /app/subscribers.xlsx

Или локально (укажи путь к БД и xlsx):
    python3 import_old_users.py subscribers.xlsx
"""
import asyncio
import aiosqlite
import sys
import re
from datetime import datetime
from openpyxl import load_workbook


DB_PATH = "data/vipsub.db"

# ID тарифа для импортированных подписок (создастся автоматически)
IMPORT_TARIFF_ID = 99


async def ensure_import_tariff(db):
    """Создать специальный тариф 'Импорт' если его нет."""
    await db.execute("""
        INSERT OR IGNORE INTO tariffs (id, name, description, days, price, sort_order, is_trial, chat_index)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (IMPORT_TARIFF_ID, "📦 Импорт (старый бот)", "Перенесено из старого бота", 30, 0, 99, 0, 0))
    await db.commit()


def parse_subscription(sub_str: str) -> tuple[datetime | None, datetime | None]:
    """
    Парсит строку вида:
    '[✅] 👥Название - 06.05.2026 15:30 - 06.06.2026 15:30'
    Возвращает (starts_at, expires_at) или (None, None) если не удалось.
    """
    if not sub_str or "[✅]" not in sub_str:
        return None, None

    # Ищем две даты в конце строки: ДД.ММ.ГГГГ ЧЧ:ММ - ДД.ММ.ГГГГ ЧЧ:ММ
    pattern = r"(\d{2}\.\d{2}\.\d{4} \d{2}:\d{2})\s*-\s*(\d{2}\.\d{2}\.\d{4} \d{2}:\d{2})\s*$"
    match = re.search(pattern, sub_str)
    if not match:
        return None, None

    try:
        starts = datetime.strptime(match.group(1), "%d.%m.%Y %H:%M")
        expires = datetime.strptime(match.group(2), "%d.%m.%Y %H:%M")
        return starts, expires
    except ValueError:
        return None, None


async def import_users(xlsx_path: str):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]
    data = rows[1:]

    print(f"📊 Найдено {len(data)} строк для импорта")
    print(f"📋 Колонки: {headers}\n")

    now = datetime.utcnow()
    imported = 0
    skipped = 0
    already_exists = 0

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await ensure_import_tariff(db)

        for row in data:
            user_id    = row[0]   # User ID
            username   = row[1]   # Username
            full_name  = row[2]   # Полное имя
            sub_str    = row[8]   # Активные подписки

            if not user_id:
                skipped += 1
                continue

            # Парсим даты подписки
            starts_at, expires_at = parse_subscription(str(sub_str) if sub_str else "")

            # Добавить пользователя
            await db.execute("""
                INSERT INTO users (user_id, username, full_name)
                VALUES (?, ?, ?)
                ON CONFLICT(user_id) DO UPDATE SET
                    username = excluded.username,
                    full_name = excluded.full_name
            """, (user_id, username, full_name or ""))

            # Проверить нет ли уже активной подписки на чат 0
            async with db.execute("""
                SELECT s.id FROM subscriptions s
                JOIN tariffs t ON t.id = s.tariff_id
                WHERE s.user_id = ? AND s.is_active = 1
                  AND (s.chat_index = 0 OR t.id = 99)
            """, (user_id,)) as cur:
                existing = await cur.fetchone()

            if existing:
                already_exists += 1
                print(f"  ⏭  {full_name} ({user_id}) — подписка уже есть, пропускаем")
                continue

            if expires_at and expires_at > now:
                is_active = 1
                status = "✅ активна"
            elif expires_at:
                is_active = 0
                status = "❌ истекла"
            else:
                is_active = 0
                starts_at = now
                expires_at = now
                status = "❓ нет данных"

            await db.execute("""
                INSERT INTO subscriptions
                    (user_id, tariff_id, starts_at, expires_at, is_active, chat_index)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                user_id,
                IMPORT_TARIFF_ID,
                (starts_at or now).isoformat(),
                (expires_at or now).isoformat(),
                is_active,
                0
            ))

            imported += 1
            print(f"  {'✅' if is_active else '⚪'} {full_name} (@{username or '—'}) | {status}"
                  + (f" до {expires_at.strftime('%d.%m.%Y')}" if expires_at and expires_at > now else ""))

        await db.commit()

    print(f"\n{'='*50}")
    print(f"✅ Импортировано:     {imported}")
    print(f"⏭  Уже были в боте:  {already_exists}")
    print(f"⚠️  Пропущено:        {skipped}")
    print(f"{'='*50}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "subscribers.xlsx"
    asyncio.run(import_users(path))
