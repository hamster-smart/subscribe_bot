"""
Импорт подписчиков второго чата (chat_index=1).

Запуск:
    docker compose cp import_chat2.py bot:/app/
    docker compose cp subscribers_chat2.xlsx bot:/app/
    docker compose exec bot python3 import_chat2.py /app/subscribers_chat2.xlsx
"""
import asyncio
import aiosqlite
import sys
import re
from datetime import datetime
from openpyxl import load_workbook

DB_PATH = "data/vipsub.db"
CHAT_INDEX = 1          # второй чат
IMPORT_TARIFF_ID = 98   # отдельный тариф для импорта чата 2


async def ensure_import_tariff(db):
    await db.execute("""
        INSERT OR IGNORE INTO tariffs (id, name, description, days, price, sort_order, is_trial, chat_index)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (IMPORT_TARIFF_ID, "📦 Импорт чат 2 (старый бот)", "Перенесено из старого бота", 30, 0, 99, 0, CHAT_INDEX))
    await db.commit()


def parse_subscription(sub_str: str):
    if not sub_str or "[✅]" not in sub_str:
        return None, None
    pattern = r"(\d{2}\.\d{2}\.\d{4} \d{2}:\d{2})\s*-\s*(\d{2}\.\d{2}\.\d{4} \d{2}:\d{2})\s*$"
    match = re.search(pattern, sub_str)
    if not match:
        return None, None
    try:
        starts = datetime.strptime(match.group(1), "%d.%m.%Y %H:%M")
        expires = datetime.strptime(match.group(2), "%d.%m.%Y %H:%M")
        return starts, expires
    except ValueError:
        return None, None


async def import_users(xlsx_path: str):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    data = rows[1:]

    print(f"📊 Найдено {len(data)} строк для импорта (Чат 2)\n")

    now = datetime.utcnow()
    imported = 0
    skipped = 0
    already_exists = 0

    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        await ensure_import_tariff(db)

        for row in data:
            user_id   = row[0]
            username  = row[1]
            full_name = row[2]
            sub_str   = row[8]

            if not user_id:
                skipped += 1
                continue

            starts_at, expires_at = parse_subscription(str(sub_str) if sub_str else "")

            # Добавить пользователя
            await db.execute("""
                INSERT INTO users (user_id, username, full_name)
                VALUES (?, ?, ?)
                ON CONFLICT(user_id) DO UPDATE SET
                    username = excluded.username,
                    full_name = excluded.full_name
            """, (user_id, username, full_name or ""))

            # Проверить — есть ли уже подписка на чат 2
            async with db.execute("""
                SELECT s.id FROM subscriptions s
                JOIN tariffs t ON t.id = s.tariff_id
                WHERE s.user_id = ? AND s.is_active = 1
                  AND (t.chat_index = 1 OR t.id = ?)
            """, (user_id, IMPORT_TARIFF_ID)) as cur:
                existing = await cur.fetchone()

            if existing:
                already_exists += 1
                print(f"  ⏭  {full_name} ({user_id}) — подписка на чат 2 уже есть")
                continue

            if expires_at and expires_at > now:
                is_active = 1
                status = f"✅ активна до {expires_at.strftime('%d.%m.%Y')}"
            elif expires_at:
                is_active = 0
                status = "❌ истекла"
            else:
                is_active = 0
                starts_at = now
                expires_at = now
                status = "❓ нет данных"

            await db.execute("""
                INSERT INTO subscriptions
                    (user_id, tariff_id, starts_at, expires_at, is_active)
                VALUES (?, ?, ?, ?, ?)
            """, (
                user_id,
                IMPORT_TARIFF_ID,
                (starts_at or now).isoformat(),
                (expires_at or now).isoformat(),
                is_active
            ))

            imported += 1
            icon = "✅" if is_active else "⚪"
            print(f"  {icon} {full_name} (@{username or '—'}) | {status}")

        await db.commit()

    print(f"\n{'='*50}")
    print(f"✅ Импортировано:     {imported}")
    print(f"⏭  Уже были:         {already_exists}")
    print(f"⚠️  Пропущено:        {skipped}")
    print(f"{'='*50}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "subscribers_chat2.xlsx"
    asyncio.run(import_users(path))
