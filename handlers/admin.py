from aiogram import Router, F, Bot
from aiogram.types import CallbackQuery, Message
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import database as db
from config import config
from keyboards.inline import (
    admin_menu_kb, admin_payment_kb, back_kb,
    admin_tariffs_kb, tariff_edit_kb, promos_kb, cancel_kb
)
from services.channel import grant_access, kick_user, mute_user
from datetime import datetime

router = Router()


def is_admin(user_id: int) -> bool:
    return user_id in config.ADMIN_IDS


CURRENCY_SYMBOLS = {"RUB": "₽", "EUR": "€", "USD": "$"}


def fmt_amount(amount, currency: str) -> str:
    symbol = CURRENCY_SYMBOLS.get(currency, currency or "₽")
    return f"{amount:.0f} {symbol}"


# ─── ADMIN MENU ──────────────────────────────────────────────────────────────

@router.message(Command("admin"))
async def cmd_admin(message: Message):
    if not is_admin(message.from_user.id):
        return
    await message.answer("<b>👑 Панель администратора</b>", reply_markup=admin_menu_kb(), parse_mode="HTML")


@router.callback_query(F.data == "admin_menu")
async def cb_admin_menu(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    await call.message.edit_text("<b>👑 Панель администратора</b>", reply_markup=admin_menu_kb(), parse_mode="HTML")


# ─── STATS ───────────────────────────────────────────────────────────────────

@router.callback_query(F.data == "admin_stats")
async def cb_admin_stats(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    import aiosqlite as aiosqlite
    from config import config as cfg
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        total_users = (await (await dbc.execute("SELECT COUNT(*) as c FROM users")).fetchone())["c"]
        pending = (await (await dbc.execute(
            "SELECT COUNT(*) as c FROM payments WHERE status='pending' AND method LIKE 'manual%'"
        )).fetchone())["c"]
       # выручка по валютам — текущий месяц
        async with dbc.execute(
            "SELECT COALESCE(p.currency, pm.currency, 'RUB') as cur, COALESCE(SUM(p.amount),0) as s "
            "FROM payments p LEFT JOIN payment_methods pm ON p.payment_method_id=pm.id "
            "WHERE p.status='confirmed' "
            "AND strftime('%Y-%m', p.confirmed_at) = strftime('%Y-%m', 'now') "
            "GROUP BY COALESCE(p.currency, pm.currency, 'RUB') ORDER BY cur"
        ) as _cur:
            revenue_rows = await _cur.fetchall()
        # выручка по валютам — сегодня
        async with dbc.execute(
            "SELECT COALESCE(p.currency, pm.currency, 'RUB') as cur, COALESCE(SUM(p.amount),0) as s "
            "FROM payments p LEFT JOIN payment_methods pm ON p.payment_method_id=pm.id "
            "WHERE p.status='confirmed' AND date(p.confirmed_at)=date('now') "
            "GROUP BY COALESCE(p.currency, pm.currency, 'RUB') ORDER BY cur"
        ) as _cur2:
            today_rows = await _cur2.fetchall()
        chat_stats = []
        for idx in range(2):
            name = cfg.get_channel_name(idx)
            row2 = await (await dbc.execute(
                "SELECT COUNT(*) as c FROM subscriptions s "
                "JOIN tariffs t ON t.id=s.tariff_id "
                "WHERE s.is_active=1 AND datetime(s.expires_at)>datetime('now') AND t.chat_index=?",
                (idx,)
            )).fetchone()
            chat_stats.append((name, row2["c"]))
    chat1_name, chat1_cnt = chat_stats[0]
    chat2_name, chat2_cnt = chat_stats[1]
    text = (
        f"📊 <b>Статистика</b>\n\n"
        f"👥 <b>Пользователей:</b> {total_users}\n"
        f"<b>{chat1_name}:</b> {chat1_cnt}\n"
        f"<b>{chat2_name}:</b> {chat2_cnt}\n\n"
        f"💰 <b>Выручка за {datetime.now().strftime('%B %Y')}:</b>\n"
        + "\n".join(
            f"  {CURRENCY_SYMBOLS.get(r['cur'], r['cur'])} {r['s']:.0f} {r['cur']}"
            for r in revenue_rows
        )
        + ("\n" if revenue_rows else "")
        + "\n📅 <b>Сегодня:</b>\n"
        + "\n".join(
            f"  {CURRENCY_SYMBOLS.get(r['cur'], r['cur'])} {r['s']:.0f} {r['cur']}"
            for r in today_rows
        )
        + ("\n" if today_rows else "нет\n")
        + f"\n⏳ <b>Ожидают:</b> {pending}"
    )
    await call.message.edit_text(text, reply_markup=back_kb("admin_menu"), parse_mode="HTML")


# ─── PENDING PAYMENTS ────────────────────────────────────────────────────────

@router.callback_query(F.data == "admin_pending")
async def cb_admin_pending(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    payments = await db.get_pending_payments()
    if not payments:
        await call.message.edit_text("✅ Нет ожидающих платежей.", reply_markup=back_kb("admin_menu"))
        return
    await call.message.edit_text(f"⏳ <b>{len(payments)} платежей</b> ожидают проверки...", parse_mode="HTML")
    for p in payments:
        currency = (p["currency"] if "currency" in p.keys() and p["currency"] else "RUB")
        amount_str = fmt_amount(p["amount"], currency)
        text = (
            f"💳 <b>Платёж #{p['id']}</b>\n"
            f"👤 {p['full_name']} "
            f"{'@' + p['username'] if p['username'] else '<code>' + str(p['user_id']) + '</code>'}\n"
            f"📦 {p['tariff_name']} · {p['days']} дн.\n"
            f"💰 {amount_str} · {p['method']}\n"
            f"🕐 {p['created_at']}"
        )
        kb = admin_payment_kb(p["id"])
        if p["screenshot_file_id"]:
            await call.bot.send_photo(
                call.from_user.id, photo=p["screenshot_file_id"],
                caption=text, reply_markup=kb, parse_mode="HTML"
            )
        else:
            await call.bot.send_message(call.from_user.id, text, reply_markup=kb, parse_mode="HTML")


@router.callback_query(F.data.startswith("admin_confirm"))
async def cb_admin_confirm(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    payment_id = int(call.data.split(":")[1])
    payment = await db.get_payment(payment_id)
    if not payment:
        await call.answer("❌ Платёж не найден.", show_alert=True)
        return
    if payment["status"] != "pending":
        await call.answer(f"Статус: {payment['status']}", show_alert=True)
        return

    await db.confirm_payment(payment_id, call.from_user.id)
    tariff = await db.get_tariff(payment["tariff_id"])
    chat_index = payment["chat_index"] if payment["chat_index"] is not None else 0
    await db.create_subscription(payment["user_id"], payment["tariff_id"], tariff["days"], chat_index)
    from services.channel import unmute_user
    await unmute_user(bot, payment["user_id"], chat_index)
    link = await grant_access(bot, payment["user_id"], chat_index)
    from datetime import datetime, timedelta
    expires = datetime.utcnow() + timedelta(days=tariff["days"])
    currency = (payment["currency"] if "currency" in payment.keys() and payment["currency"] else (tariff["currency"] if "currency" in tariff.keys() and tariff["currency"] else "RUB"))
    amount_str = fmt_amount(payment["amount"], currency)
    try:
        await bot.send_message(
            payment["user_id"],
            f"✅ <b>Оплата подтверждена!</b>\n"
            f"📦 <b>{tariff['name']}</b>\n"
            f"💰 {amount_str}\n"
            f"🔗 {link}\n"
            f"📅 До: <b>{expires.strftime('%d.%m.%Y')}</b>\n\n"
            f"👉 Перейдите по ссылке в чат и нажмите кнопку «Подать заявку на вступление...».",
            parse_mode="HTML"
        )
    except Exception:
        pass

    suffix = " ✅ <b>Подтверждён</b>"
    if call.message.caption:
        await call.message.edit_caption(call.message.caption + suffix, parse_mode="HTML")
    else:
        await call.message.edit_text(call.message.text + suffix, parse_mode="HTML")
    await call.answer("✅ Подтверждено!", show_alert=False)

      # ─── Уведомление администраторам о ручном подтверждении ──
    if not tariff.get("is_trial"):
        import aiosqlite
        async with aiosqlite.connect(config.DB_PATH) as dbc:
            dbc.row_factory = aiosqlite.Row
            async with dbc.execute(
                "SELECT username, full_name FROM users WHERE user_id = ?",
                (payment["user_id"],)
            ) as cur:
                user_row = await cur.fetchone()

        username = user_row["username"] if user_row and user_row["username"] else None
        full_name = user_row["full_name"] if user_row else str(payment["user_id"])
        currency_symbols = {"RUB": "₽", "EUR": "€", "USD": "$"}
        symbol = currency_symbols.get(currency, currency)
        chat_name = config.get_channel_name(chat_index)

        admin_text = (
            f"🆕 <b>Новая подписка!</b>\n\n"
            f"👤 {full_name}"
            + (f" (@{username})" if username else "")
            + f"\n🆔 <code>{payment['user_id']}</code>\n"
            f"📦 Тариф: {tariff['name']}\n"
            f"📺 Канал: {chat_name}\n"
            f"💵 Сумма: <b>{payment['amount']:.0f} {symbol}</b>\n"
            f"💳 Метод: Ручной перевод\n"
            f"📅 До: {expires.strftime('%d.%m.%Y')}"
        )
        for admin_id in config.ADMIN_IDS:
            try:
                await bot.send_message(admin_id, admin_text, parse_mode="HTML")
            except Exception:
                pass


@router.callback_query(F.data.startswith("admin_reject"))
async def cb_admin_reject(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    payment_id = int(call.data.split(":")[1])
    payment = await db.get_payment(payment_id)
    if not payment or payment["status"] != "pending":
        await call.answer("❌", show_alert=True)
        return
    await db.reject_payment(payment_id, call.from_user.id)
    try:
        await bot.send_message(
            payment["user_id"],
            "❌ Ваш платёж отклонён.\n"
            "Свяжитесь с поддержкой, если это ошибка."
        )
    except Exception:
        pass
    await call.answer("❌ Отклонён", show_alert=False)
    suffix = " ❌ <b>Отклонён</b>"
    if call.message.caption:
        await call.message.edit_caption(call.message.caption + suffix, parse_mode="HTML")
    else:
        await call.message.edit_text(call.message.text + suffix, parse_mode="HTML")


# ─── SUBSCRIBERS ─────────────────────────────────────────────────────────────

@router.callback_query(F.data == "admin_subs")
async def cb_admin_subs(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    from config import config as cfg
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text=f"📋 {cfg.CHANNEL_1_NAME}", callback_data="admin_subs_export:0"))
    builder.row(InlineKeyboardButton(text=f"📋 {cfg.CHANNEL_2_NAME}", callback_data="admin_subs_export:1"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="admin_menu"))
    await call.message.edit_text("<b>📋 Экспорт подписчиков</b>", reply_markup=builder.as_markup(), parse_mode="HTML")


@router.callback_query(F.data.startswith("admin_subs_export"))
async def cb_admin_subs_export(call: CallbackQuery, bot: Bot):
    chat_index = int(call.data.split(":")[1])
    if not is_admin(call.from_user.id):
        return
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute(
            "SELECT u.user_id, u.username, u.full_name, s.starts_at, s.expires_at, "
            "t.name as tariff_name, s.is_active "
            "FROM subscriptions s "
            "JOIN users u ON u.user_id=s.user_id "
            "JOIN tariffs t ON t.id=s.tariff_id "
            "WHERE s.is_active=1 AND datetime(s.expires_at)>datetime('now') AND s.chat_index=? "
            "ORDER BY s.expires_at ASC",
            (chat_index,)
        ) as cur:
            subs = await cur.fetchall()
    from config import config as cfg
    chat_name = cfg.get_channel_name(chat_index)
    if not subs:
        await call.message.edit_text(f"{chat_name}: нет активных подписчиков.", reply_markup=back_kb("admin_subs"))
        return
    import io
    from datetime import datetime, datetime as dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from aiogram.types import BufferedInputFile
    wb = Workbook()
    ws = wb.active
    ws.title = f"Чат {chat_index + 1}"
    headers = ["User ID", "Username", "Имя", "Тариф", "С", "До", "Дней осталось"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E86AB")
        cell.alignment = Alignment(horizontal="center")
    now = datetime.utcnow()
    for s in subs:
        exp = datetime.fromisoformat(s["expires_at"])
        start = datetime.fromisoformat(s["starts_at"])
        ws.append([
            s["user_id"],
            f"@{s['username']}" if s["username"] else "",
            s["full_name"] or "",
            s["tariff_name"],
            start.strftime("%d.%m.%Y"),
            exp.strftime("%d.%m.%Y"),
            (exp - now).days,
        ])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 10), 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"subs_chat{chat_index + 1}_{dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await call.message.edit_text(f"<b>{chat_name}</b> — {len(subs)} подписчиков. Генерирую...", parse_mode="HTML")
    await bot.send_document(
        call.from_user.id,
        document=BufferedInputFile(buf.read(), filename=filename),
        caption=f"{chat_name} · {dt.now().strftime('%d.%m.%Y %H:%M')} · {len(subs)} чел."
    )
    await call.message.edit_text(
        f"✅ {len(subs)} подписчиков {chat_name}.",
        reply_markup=back_kb("admin_subs"), parse_mode="HTML"
    )


# ─── USER LOOKUP ─────────────────────────────────────────────────────────────

class UserLookupState(StatesGroup):
    waiting_query = State()


@router.callback_query(F.data == "admin_find_user")
async def cb_admin_find_user(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    await state.set_state(UserLookupState.waiting_query)
    await call.message.edit_text(
        "<b>🔍 Поиск пользователя</b>\n\nВведи Telegram ID или username:",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )


@router.message(UserLookupState.waiting_query)
async def handle_user_lookup(message: Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(None)
    query = message.text.strip().lstrip("@")
    import aiosqlite as aiosqlite
    from datetime import datetime
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        if query.isdigit():
            sql = "SELECT * FROM users WHERE user_id=?"
            params = (int(query),)
        else:
            sql = "SELECT * FROM users WHERE username LIKE ? OR full_name LIKE ? LIMIT 10"
            params = (f"%{query}%", f"%{query}%")
        async with dbc.execute(sql, params) as cur:
            users = await cur.fetchall()
    if not users:
        await message.answer("❌ Пользователь не найден.", reply_markup=back_kb("admin_menu"))
        return
    if len(users) > 1:
        builder = InlineKeyboardBuilder()
        for u in users:
            uname_str = f" @{u['username']}" if u['username'] else ''
            banned = ' 🚫' if u['is_banned'] else ''
            label = f"{u['full_name'] or '—'}{uname_str}{banned} [{u['user_id']}]"
            builder.row(InlineKeyboardButton(text=label[:80], callback_data=f"admin_user_info:{u['user_id']}"))
        await message.answer(f"Найдено {len(users)} пользователей:", reply_markup=builder.as_markup())
        return
    await show_user_info(message, users[0]["user_id"], bot)


@router.callback_query(F.data.startswith("admin_user_info"))
async def cb_admin_user_info(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    user_id = int(call.data.split(":")[1])
    await show_user_info(call.message, user_id, bot, edit=True)


async def show_user_info(message, user_id: int, bot: Bot, edit: bool = False):
    import aiosqlite as aiosqlite
    from datetime import datetime
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute("SELECT * FROM users WHERE user_id=?", (user_id,)) as cur:
            user = await cur.fetchone()
        async with dbc.execute(
            "SELECT s.*, t.name as tariff_name FROM subscriptions s "
            "JOIN tariffs t ON t.id=s.tariff_id "
            "WHERE s.user_id=? ORDER BY s.created_at DESC LIMIT 5",
            (user_id,)
        ) as cur:
            subs = await cur.fetchall()
        async with dbc.execute(
            "SELECT * FROM payments WHERE user_id=? ORDER BY created_at DESC LIMIT 5",
            (user_id,)
        ) as cur:
            payments = await cur.fetchall()

    if not user:
        text = "❌ Пользователь не найден."
    else:
        now = datetime.utcnow()
        uname = f"@{user['username']}" if user["username"] else ""
        joined = datetime.fromisoformat(user["joined_at"]).strftime("%d.%m.%Y") if user["joined_at"] else "—"
        banned_mark = " 🚫" if user["is_banned"] else ""
        text = (
            f"👤 <b>{user['full_name'] or '—'}</b>{banned_mark} {uname}\n"
            f"🆔 <code>{user_id}</code>\n"
            f"📅 Регистрация: {joined}\n"
        )
        if subs:
            text += "\n<b>📋 Подписки:</b>\n"
            for s in subs:
                exp = datetime.fromisoformat(s["expires_at"])
                if s["is_active"] and exp > now:
                    days_left = (exp - now).days
                    status = f"✅ активна, {days_left} дн."
                elif s["is_active"]:
                    status = "⏰ истекла"
                else:
                    status = "❌ отозвана"
                text += f"• {s['tariff_name']} — {status} · до {exp.strftime('%d.%m.%Y')}\n"
        else:
            text += "\n_Нет подписок_"
        if payments:
            text += "\n<b>💳 Платежи:</b>\n"
            for p in payments:
                status_map = {"confirmed": "✅", "pending": "⏳", "rejected": "❌"}
                icon = status_map.get(p["status"], "?")
                created = datetime.fromisoformat(p["created_at"]).strftime("%d.%m.%Y")
                currency = (p["currency"] if "currency" in p.keys() and p["currency"] else "RUB")
                amount_str = fmt_amount(p["amount"], currency)
                text += f"{icon} {amount_str} · {p['method']} · {created}\n"

    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🎁 Выдать подписку", callback_data=f"admin_grant:{user_id}"))
    builder.row(InlineKeyboardButton(text="➕ +30 дней", callback_data=f"admin_add30:{user_id}"))  # ← добавить
    builder.row(InlineKeyboardButton(text="🚫 Отозвать доступ", callback_data=f"admin_revoke:{user_id}"))
    builder.row(InlineKeyboardButton(text="🦵 Кик из канала", callback_data=f"admin_kick_user:{user_id}"))
    builder.row(InlineKeyboardButton(text="🔄 Сбросить триал", callback_data=f"admin_reset_trial:{user_id}"))
    if user and user["is_banned"]:
        builder.row(InlineKeyboardButton(text="✅ Разбанить в боте", callback_data=f"admin_unban_user:{user_id}"))
    else:
        builder.row(InlineKeyboardButton(text="🚫 Забанить в боте", callback_data=f"admin_ban_user:{user_id}"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="admin_menu"))

    if edit:
        await message.edit_text(text, reply_markup=builder.as_markup(), parse_mode="HTML")
    else:
        await message.answer(text, reply_markup=builder.as_markup(), parse_mode="HTML")


# ─── CHANNEL CHOICE KB ───────────────────────────────────────────────────────

def channel_choice_kb(action: str, user_id: int):
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    from config import config as cfg
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text=cfg.get_channel_name(0), callback_data=f"{action}:{user_id}:0"))
    builder.row(InlineKeyboardButton(text=cfg.get_channel_name(1), callback_data=f"{action}:{user_id}:1"))
    builder.row(InlineKeyboardButton(text="🔀 Оба канала", callback_data=f"{action}:{user_id}:all"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data=f"admin_user_info:{user_id}"))
    return builder.as_markup()


# ─── REVOKE / KICK / RESET TRIAL ─────────────────────────────────────────────

@router.callback_query(F.data.startswith("admin_revoke"))
async def cb_admin_revoke(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    parts = call.data.split(":")
    user_id = int(parts[1])
    if len(parts) < 3:
        await call.message.edit_text("Выбери канал:", reply_markup=channel_choice_kb("admin_revoke", user_id))
        return
    chat_index = parts[2]
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        if chat_index == "all":
            await dbc.execute("UPDATE subscriptions SET is_active=0 WHERE user_id=? AND is_active=1", (user_id,))
        else:
            await dbc.execute(
                "UPDATE subscriptions SET is_active=0 WHERE user_id=? AND is_active=1 "
                "AND tariff_id IN (SELECT id FROM tariffs WHERE chat_index=?)",
                (user_id, int(chat_index))
            )
        await dbc.commit()
    label = "Все каналы" if chat_index == "all" else config.get_channel_name(int(chat_index))
    await call.answer(f"✅ Доступ отозван: {label}", show_alert=True)
    await show_user_info(call.message, user_id, bot, edit=True)


@router.callback_query(F.data.startswith("admin_kick_user"))
async def cb_admin_kick_user(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    parts = call.data.split(":")
    user_id = int(parts[1])
    if len(parts) < 3:
        await call.message.edit_text("Выбери канал:", reply_markup=channel_choice_kb("admin_kick_user", user_id))
        return
    chat_index = parts[2]
    import aiosqlite as aiosqlite
    if chat_index == "all":
        await kick_user(bot, user_id, 0)
        await kick_user(bot, user_id, 1)
        async with aiosqlite.connect(config.DB_PATH) as dbc:
            await dbc.execute("UPDATE subscriptions SET is_active=0 WHERE user_id=? AND is_active=1", (user_id,))
            await dbc.commit()
        label = "Оба канала"
    else:
        idx = int(chat_index)
        await kick_user(bot, user_id, idx)
        async with aiosqlite.connect(config.DB_PATH) as dbc:
            await dbc.execute(
                "UPDATE subscriptions SET is_active=0 WHERE user_id=? AND is_active=1 "
                "AND tariff_id IN (SELECT id FROM tariffs WHERE chat_index=?)",
                (user_id, idx)
            )
            await dbc.commit()
        label = config.get_channel_name(idx)
    await call.answer(f"🦵 Кик: {label}", show_alert=True)
    await show_user_info(call.message, user_id, bot, edit=True)


@router.callback_query(F.data.startswith("admin_reset_trial"))
async def cb_admin_reset_trial(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    parts = call.data.split(":")
    user_id = int(parts[1])
    if len(parts) < 3:
        await call.message.edit_text("Выбери канал:", reply_markup=channel_choice_kb("admin_reset_trial", user_id))
        return
    chat_index = parts[2]
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        if chat_index == "all":
            await dbc.execute(
                "DELETE FROM subscriptions WHERE user_id=? AND tariff_id IN (SELECT id FROM tariffs WHERE is_trial=1)",
                (user_id,)
            )
            label = "Все каналы"
        else:
            await dbc.execute(
                "DELETE FROM subscriptions WHERE user_id=? AND tariff_id IN "
                "(SELECT id FROM tariffs WHERE is_trial=1 AND chat_index=?)",
                (user_id, int(chat_index))
            )
            label = config.get_channel_name(int(chat_index))
        await dbc.commit()
    await call.answer(f"🔄 Триал сброшен: {label}", show_alert=True)
    await show_user_info(call.message, user_id, bot, edit=True)


# ─── BAN / UNBAN ─────────────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("admin_ban_user"))
async def cb_admin_ban_user(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    user_id = int(call.data.split(":")[1])
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        await dbc.execute("UPDATE users SET is_banned=1 WHERE user_id=?", (user_id,))
        await dbc.commit()
    await call.answer("🚫 Забанен в боте", show_alert=True)
    await show_user_info(call.message, user_id, bot, edit=True)


@router.callback_query(F.data.startswith("admin_unban_user"))
async def cb_admin_unban_user(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    user_id = int(call.data.split(":")[1])
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        await dbc.execute("UPDATE users SET is_banned=0 WHERE user_id=?", (user_id,))
        await dbc.commit()
    await call.answer("✅ Разбанен", show_alert=False)
    await show_user_info(call.message, user_id, bot, edit=True)

@router.callback_query(F.data.startswith("admin_add30"))
async def cb_admin_add30(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    parts = call.data.split(":")
    user_id = int(parts[1])
    if len(parts) < 3:
        # Показываем выбор чата (без кнопки "все")
        from aiogram.utils.keyboard import InlineKeyboardBuilder
        from aiogram.types import InlineKeyboardButton
        from config import config as cfg
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text=cfg.get_channel_name(0), callback_data=f"admin_add30:{user_id}:0"))
        builder.row(InlineKeyboardButton(text=cfg.get_channel_name(1), callback_data=f"admin_add30:{user_id}:1"))
        builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data=f"admin_userinfo:{user_id}"))
        await call.message.edit_text("Для какого чата добавить 30 дней?", reply_markup=builder.as_markup())
        return
    chat_index = int(parts[2])
    import aiosqlite
    from datetime import datetime, timedelta
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute(
            "SELECT * FROM subscriptions WHERE user_id=? AND is_active=1 AND chat_index=? ORDER BY expires_at DESC LIMIT 1",
            (user_id, chat_index)
        ) as cur:
            sub = await cur.fetchone()
        if sub:
            new_exp = datetime.fromisoformat(sub["expires_at"]) + timedelta(days=30)
            await dbc.execute("UPDATE subscriptions SET expires_at=? WHERE id=?", (new_exp.isoformat(), sub["id"]))
        else:
            async with dbc.execute(
                "SELECT * FROM tariffs WHERE is_active=1 AND is_trial=0 AND chat_index=? ORDER BY sort_order LIMIT 1",
                (chat_index,)
            ) as cur2:
                base = await cur2.fetchone()
            tariff_id = base["id"] if base else 2
            now = datetime.utcnow()
            new_exp = now + timedelta(days=30)
            await dbc.execute(
                "INSERT INTO subscriptions (user_id, tariff_id, starts_at, expires_at, chat_index) VALUES (?,?,?,?,?)",
                (user_id, tariff_id, now.isoformat(), new_exp.isoformat(), chat_index)
            )
        await dbc.commit()
    chat_name = config.get_channel_name(chat_index)
    try:
        await bot.send_message(user_id, f"<b>+30 дней подписки!</b>\nЧат: {chat_name}\nДействует до: <b>{new_exp.strftime('%d.%m.%Y')}</b>", parse_mode="HTML")
    except Exception:
        pass
    await call.answer(f"+30 дней ({chat_name})!", show_alert=False)
    await show_user_info(call.message, user_id, bot, edit=True)


# ─── GRANT SUBSCRIPTION ──────────────────────────────────────────────────────

@router.callback_query(F.data.startswith("admin_grant:"))
async def cb_admin_grant(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    user_id = int(call.data.split(":")[1])
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute(
            "SELECT * FROM tariffs WHERE is_active=1 AND is_trial=0 ORDER BY chat_index, sort_order, id"
        ) as cur:
            tariffs = await cur.fetchall()
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    for t in tariffs:
        chat_name = config.get_channel_name(t["chat_index"] or 0)
        currency = t["currency"] or "RUB"
        builder.row(InlineKeyboardButton(
            text=f"{t['name']} {chat_name} · {t['days']} дн. · {fmt_amount(t['price'], currency)}",
            callback_data=f"admin_grant_tariff:{user_id}:{t['id']}"
        ))
    builder.row(InlineKeyboardButton(text="✏️ Произвольный срок", callback_data=f"admin_grant_custom:{user_id}"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data=f"admin_user_info:{user_id}"))
    await call.message.edit_text("Выбери тариф:", reply_markup=builder.as_markup())


@router.callback_query(F.data.startswith("admin_grant_tariff"))
async def cb_admin_grant_tariff(call: CallbackQuery, bot: Bot):
    if not is_admin(call.from_user.id):
        return
    _, user_id, tariff_id = call.data.split(":")
    user_id, tariff_id = int(user_id), int(tariff_id)
    tariff = await db.get_tariff(tariff_id)
    chat_index = tariff["chat_index"] or 0
    await db.create_subscription(user_id, tariff_id, tariff["days"], chat_index)
    from services.channel import grant_access
    link = await grant_access(bot, user_id, chat_index)
    from datetime import datetime, timedelta
    expires = datetime.utcnow() + timedelta(days=tariff["days"])
    chat_name = config.get_channel_name(chat_index)
    currency = (tariff["currency"] if "currency" in tariff.keys() and tariff["currency"] else "RUB")
    try:
        await bot.send_message(
            user_id,
            f"🎁 <b>Подписка выдана!</b>\n"
            f"📦 {tariff['name']} · {chat_name}\n"
            f"🔗 {link}\n"
            f"📅 До: <b>{expires.strftime('%d.%m.%Y')}</b>",
            parse_mode="HTML"
        )
    except Exception:
        pass
    await call.answer(f"✅ Выдано {tariff['days']} дн.")
    await show_user_info(call.message, user_id, bot, edit=True)


class GrantCustomState(StatesGroup):
    days = State()


@router.callback_query(F.data.startswith("admin_grant_custom"))
async def cb_admin_grant_custom(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    parts = call.data.split(":")
    user_id = int(parts[1])
    if len(parts) < 3:
        # Выбор чата
        from aiogram.utils.keyboard import InlineKeyboardBuilder
        from aiogram.types import InlineKeyboardButton
        from config import config as cfg
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text=cfg.get_channel_name(0), callback_data=f"admin_grant_custom:{user_id}:0"))
        builder.row(InlineKeyboardButton(text=cfg.get_channel_name(1), callback_data=f"admin_grant_custom:{user_id}:1"))
        builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data=f"admin_userinfo:{user_id}"))
        await call.message.edit_text("Для какого чата выдать подписку?", reply_markup=builder.as_markup())
        return
    chat_index = int(parts[2])
    await state.update_data(grant_user_id=user_id, grant_chat_index=chat_index)
    await state.set_state(GrantCustomState.days)
    await call.message.edit_text("Введите количество дней:", reply_markup=back_kb(f"admin_grant:{user_id}"))


@router.message(GrantCustomState.days)
async def handle_grant_custom_days(message: Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id):
        return
    try:
        days = int(message.text.strip())
        if days < 1:
            raise ValueError
    except ValueError:
        await message.answer("Введите целое число от 1.")
        return
    data = await state.get_data()
    user_id = data["grant_user_id"]
    chat_index = data.get("grant_chat_index", 0)
    await state.set_state(None)
    tariff_id = 106 if chat_index == 0 else 108
    await db.create_subscription(user_id, tariff_id, days, chat_index)
    from services.channel import grant_access
    link = await grant_access(bot, user_id, chat_index)
    from datetime import datetime, timedelta
    expires = datetime.utcnow() + timedelta(days=days)
    chat_name = config.get_channel_name(chat_index)
    try:
        await bot.send_message(
            user_id,
            f"<b>Подписка на {days} дней активирована!</b>\n"
            f"Чат: {chat_name}\n"
            f"Действует до: <b>{expires.strftime('%d.%m.%Y')}</b>\n"
            f"{link}",
            parse_mode="HTML"
        )
    except Exception:
        pass
    await message.answer(f"✅ Выдано {days} дней ({chat_name})!")
    await show_user_info(message, user_id, bot, edit=False)

# ─── BROADCAST ───────────────────────────────────────────────────────────────

class BroadcastState(StatesGroup):
    waiting_message = State()


@router.callback_query(F.data == "admin_broadcast")
async def cb_admin_broadcast(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    await state.set_state(BroadcastState.waiting_message)
    await call.message.edit_text(
        "<b>📢 Рассылка</b>\n\nОтправь сообщение (текст, фото, видео). Поддерживается HTML-форматирование.",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )


@router.message(BroadcastState.waiting_message)
async def handle_broadcast(message: Message, state: FSMContext, bot: Bot):
    if not is_admin(message.from_user.id):
        return
    await state.set_state(None)
    users = await db.get_all_users()
    sent = 0
    failed = 0
    status_msg = await message.answer(f"📢 Рассылка... 0/{len(users)}")
    for i, user in enumerate(users):
        try:
            if message.photo:
                await bot.send_photo(user["user_id"], message.photo[-1].file_id, caption=message.caption, parse_mode="HTML")
            elif message.video:
                await bot.send_video(user["user_id"], message.video.file_id, caption=message.caption, parse_mode="HTML")
            else:
                await bot.send_message(user["user_id"], message.text, parse_mode="HTML")
            sent += 1
        except Exception:
            failed += 1
        if (i + 1) % 20 == 0:
            try:
                await status_msg.edit_text(f"📢 {sent}/{len(users)}...")
            except Exception:
                pass
        import asyncio
        await asyncio.sleep(0.05)
    await status_msg.edit_text(
        f"✅ Рассылка завершена!\n📨 Отправлено: {sent}\n❌ Ошибок: {failed}",
        reply_markup=back_kb("admin_menu")
    )


# ─── PROMO CODES ─────────────────────────────────────────────────────────────

class PromoCreateState(StatesGroup):
    code = State()
    discount = State()
    tariff = State()
    uses_total = State()
    uses_per_user = State()


async def show_promo_list(target, edit: bool = True):
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute("SELECT * FROM promo_codes ORDER BY created_at DESC LIMIT 20") as cur:
            promos = await cur.fetchall()
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    if promos:
        for p in promos:
            active = "✅" if p["is_active"] else "❌"
            disc = f" -{p['discount_pct']}%" if p["discount_pct"] else " 🆓"
            builder.row(InlineKeyboardButton(
                text=f"{active} {p['code']}{disc}",
                callback_data=f"promo_card:{p['id']}"
            ))
        text = "<b>🎟 Промокоды</b>"
    else:
        text = "Промокодов нет."
    builder.row(InlineKeyboardButton(text="➕ Создать промокод", callback_data="create_promo"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="admin_menu"))
    if edit:
        await target.edit_text(text, reply_markup=builder.as_markup(), parse_mode="HTML")
    else:
        await target.answer(text, reply_markup=builder.as_markup(), parse_mode="HTML")


async def show_promo_card(target, promo_id: int, edit: bool = True):
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute("SELECT * FROM promo_codes WHERE id=?", (promo_id,)) as cur:
            p = await cur.fetchone()
    if not p:
        return
    disc = f"-{p['discount_pct']}%" if p["discount_pct"] else "100% (бесплатно)"
    uses = f"{p['uses_left']} ост." if p["uses_left"] != -1 else "∞"
    per_user = str(p["max_uses_per_user"]) if p["max_uses_per_user"] else "1"
    active = "✅ Активен" if p["is_active"] else "❌ Отключён"
    bot_username = await db.get_setting("bot_username", "BOT_USERNAME")
    tariff_info = ""
    if p["tariff_id"]:
        tariff = await db.get_tariff(p["tariff_id"])
        if tariff:
            chat_name = config.get_channel_name(p["chat_index"] or 0)
            tariff_info = f"\n📦 Тариф: {tariff['name']} {chat_name}"
    text = (
        f"🎟 <b>{p['code']}</b>\n"
        f"💰 Скидка: <b>{disc}</b>{tariff_info}\n"
        f"🔢 Использований: <b>{uses}</b> · на юзера: <b>{per_user}</b>\n"
        f"📊 {active}\n\n"
        f"🔗 <code>https://t.me/{bot_username}?start={p['code']}</code>"
    )
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    toggle = "❌ Отключить" if p["is_active"] else "✅ Включить"
    builder.row(InlineKeyboardButton(text=toggle, callback_data=f"promo_toggle:{promo_id}"))
    builder.row(InlineKeyboardButton(text="🗑 Удалить", callback_data=f"promo_delete_confirm:{promo_id}"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="admin_promos"))
    if edit:
        await target.edit_text(text, reply_markup=builder.as_markup(), parse_mode="HTML")
    else:
        await target.answer(text, reply_markup=builder.as_markup(), parse_mode="HTML")


@router.callback_query(F.data == "admin_promos")
async def cb_admin_promos(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    await show_promo_list(call.message, edit=True)


@router.callback_query(F.data.startswith("promo_card"))
async def cb_promo_card(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    promo_id = int(call.data.split(":")[1])
    await show_promo_card(call.message, promo_id, edit=True)


@router.callback_query(F.data.startswith("promo_toggle"))
async def cb_promo_toggle(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    promo_id = int(call.data.split(":")[1])
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        await dbc.execute(
            "UPDATE promo_codes SET is_active=CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id=?",
            (promo_id,)
        )
        await dbc.commit()
    await show_promo_card(call.message, promo_id, edit=True)


@router.callback_query(F.data.startswith("promo_delete_confirm"))
async def cb_promo_delete_confirm(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    promo_id = int(call.data.split(":")[1])
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"promo_delete:{promo_id}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data=f"promo_card:{promo_id}")
    )
    await call.message.edit_text("Удалить промокод?", reply_markup=builder.as_markup())


@router.callback_query(F.data.startswith("promo_delete:"))
async def cb_promo_delete(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    promo_id = int(call.data.split(":")[1])
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        await dbc.execute("DELETE FROM promo_codes WHERE id=?", (promo_id,))
        await dbc.commit()
    await call.answer("🗑 Удалён")
    await show_promo_list(call.message, edit=True)


@router.callback_query(F.data == "create_promo")
async def cb_create_promo(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    await state.set_state(PromoCreateState.code)
    await call.message.edit_text(
        "<b>1/5 · Введи код промокода</b>\n\nТолько буквы и цифры, до 15 символов.\n"
        "<i>Пример: SALE30, VIP2024</i>",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )


@router.message(PromoCreateState.code)
async def promo_set_code(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    code = message.text.strip().upper()
    if not code.isalnum():
        await message.answer("❌ Только буквы и цифры.")
        return
    await state.update_data(code=code)
    await state.set_state(PromoCreateState.discount)
    await message.answer(
        "2/5 · <b>Скидка (%)</b>\n\n"
        "<b>100</b> = бесплатно\n<b>50</b> = скидка 50%\n<b>0</b> = без скидки (просто ссылка)",
        parse_mode="HTML"
    )


@router.message(PromoCreateState.discount)
async def promo_set_discount(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        pct = int(message.text.strip().replace("%", ""))
        if not (0 <= pct <= 100):
            raise ValueError
    except ValueError:
        await message.answer("❌ Введи число от 0 до 100.")
        return
    await state.update_data(disc_pct=pct)
    await state.set_state(PromoCreateState.tariff)
    import aiosqlite as ai
    async with ai.connect(config.DB_PATH) as dbc:
        dbc.row_factory = ai.Row
        async with dbc.execute(
            "SELECT * FROM tariffs WHERE is_active=1 AND is_trial=0 ORDER BY chat_index, sort_order"
        ) as cur:
            all_tariffs = await cur.fetchall()
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    for t in all_tariffs:
        chat_name = config.get_channel_name(t["chat_index"] or 0)
        currency = t["currency"] or "RUB"
        builder.row(InlineKeyboardButton(
            text=f"{t['name']} {chat_name} · {fmt_amount(t['price'], currency)}",
            callback_data=f"promo_pick_tariff:{t['id']}:{t['chat_index'] or 0}"
        ))
    builder.row(InlineKeyboardButton(text="🔀 Любой тариф", callback_data="promo_pick_tariff:0:0"))
    await message.answer("3/5 · <b>Привязать к тарифу?</b>", reply_markup=builder.as_markup(), parse_mode="HTML")


@router.callback_query(F.data.startswith("promo_pick_tariff"), PromoCreateState.tariff)
async def promo_pick_tariff(call: CallbackQuery, state: FSMContext):
    parts = call.data.split(":")
    tariff_id = int(parts[1])
    chat_index = int(parts[2])
    await state.update_data(
        promo_tariff_id=tariff_id if tariff_id != 0 else None,
        promo_chat_index=chat_index if tariff_id != 0 else None
    )
    await state.set_state(PromoCreateState.uses_total)
    await call.message.edit_text(
        "4/5 · <b>Сколько раз можно использовать?</b>\n\n<b>0</b> = неограничено",
        parse_mode="HTML"
    )


@router.message(PromoCreateState.uses_total)
async def promo_set_uses_total(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        uses = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введи число.")
        return
    await state.update_data(uses_total=uses)
    await state.set_state(PromoCreateState.uses_per_user)
    await message.answer("5/5 · <b>Макс. использований на одного юзера?</b> (обычно <b>1</b>)", parse_mode="HTML")


@router.message(PromoCreateState.uses_per_user)
async def promo_set_uses_per_user(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        per_user = int(message.text.strip())
        if per_user < 1:
            raise ValueError
    except ValueError:
        await message.answer("❌ Минимум 1.")
        return
    data = await state.get_data()
    await state.set_state(None)
    uses_left = -1 if data["uses_total"] == 0 else data["uses_total"]
    promo_id = await db.create_promo(
        code=data["code"],
        discount_pct=data.get("disc_pct", 0),
        uses_left=uses_left,
        tariff_id=data.get("promo_tariff_id"),
        chat_index=data.get("promo_chat_index"),
        max_uses_per_user=per_user
    )
    disc = f"{data['disc_pct']}%" if data.get("disc_pct") else "бесплатно"
    uses_str = "∞" if uses_left == -1 else str(uses_left)
    bot_username = await db.get_setting("bot_username", "BOT_USERNAME")
    await message.answer(
        f"✅ <b>Промокод создан!</b>\n"
        f"🎟 <code>{data['code']}</code>\n"
        f"💰 Скидка: {disc}\n"
        f"🔢 Использований: {uses_str} · на юзера: {per_user}\n\n"
        f"🔗 <code>https://t.me/{bot_username}?start={data['code']}</code>",
        reply_markup=back_kb("admin_promos"), parse_mode="HTML"
    )


# ─── TARIFFS ─────────────────────────────────────────────────────────────────

class TariffEditState(StatesGroup):
    field = State()
    value = State()


class TariffAddState(StatesGroup):
    name = State()
    description = State()
    days = State()
    price = State()
    currency = State()


def tariff_card_kb(tariff_id: int, is_active: int):
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    toggle_text = "❌ Отключить" if is_active else "✅ Включить"
    builder.row(InlineKeyboardButton(text=toggle_text, callback_data=f"tariff_toggle:{tariff_id}"))
    builder.row(InlineKeyboardButton(text="✏️ Редактировать", callback_data=f"tariff_edit_menu:{tariff_id}"))
    builder.row(InlineKeyboardButton(text="🗑 Удалить", callback_data=f"tariff_delete_confirm:{tariff_id}"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="admin_tariffs"))
    return builder.as_markup()


def tariff_edit_fields_kb(tariff_id: int):
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="📝 Название", callback_data=f"tariff_edit_field:{tariff_id}:name"))
    builder.row(InlineKeyboardButton(text="📄 Описание", callback_data=f"tariff_edit_field:{tariff_id}:description"))
    builder.row(InlineKeyboardButton(text="📅 Дней", callback_data=f"tariff_edit_field:{tariff_id}:days"))
    builder.row(InlineKeyboardButton(text="💰 Цена", callback_data=f"tariff_edit_field:{tariff_id}:price"))
    builder.row(InlineKeyboardButton(text="💱 Валюта", callback_data=f"tariff_edit_currency:{tariff_id}"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data=f"admin_edit_tariff:{tariff_id}"))
    return builder.as_markup()


async def show_tariff_card(target, tariff_id: int, edit: bool = True):
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute("SELECT * FROM tariffs WHERE id=?", (tariff_id,)) as cur:
            t = await cur.fetchone()
    if not t:
        return
    status = "✅ Активен" if t["is_active"] else "❌ Отключён"
    trial = " 🆓 Триал" if t["is_trial"] else ""
    currency = t["currency"] or "RUB"
    text = (
        f"📦 <b>{t['name']}</b>{trial}\n"
        f"{t['description'] or ''}\n\n"
        f"📅 Дней: <b>{t['days']}</b>\n"
        f"💰 Цена: <b>{fmt_amount(t['price'], currency)}</b>\n"
        f"📊 {status}"
    )
    kb = tariff_card_kb(tariff_id, t["is_active"])
    if edit:
        await target.edit_text(text, reply_markup=kb, parse_mode="HTML")
    else:
        await target.answer(text, reply_markup=kb, parse_mode="HTML")


@router.callback_query(F.data == "admin_tariffs")
async def cb_admin_tariffs(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute("SELECT * FROM tariffs ORDER BY sort_order, id") as cur:
            tariffs = await cur.fetchall()
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    for t in tariffs:
        status = "✅" if t["is_active"] else "❌"
        trial = " 🆓" if t["is_trial"] else ""
        currency = t["currency"] or "RUB"
        builder.row(InlineKeyboardButton(
            text=f"{status}{trial} {t['name']} · {fmt_amount(t['price'], currency)}",
            callback_data=f"admin_edit_tariff:{t['id']}"
        ))
    builder.row(InlineKeyboardButton(text="➕ Добавить тариф", callback_data="admin_add_tariff"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="admin_settings"))
    await call.message.edit_text("<b>📦 Тарифы</b>", reply_markup=builder.as_markup(), parse_mode="HTML")


@router.callback_query(F.data.startswith("admin_edit_tariff"))
async def cb_edit_tariff(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    tariff_id = int(call.data.split(":")[1])
    await show_tariff_card(call.message, tariff_id, edit=True)


@router.callback_query(F.data.startswith("tariff_toggle"))
async def cb_tariff_toggle(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    tariff_id = int(call.data.split(":")[1])
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute("SELECT is_active FROM tariffs WHERE id=?", (tariff_id,)) as cur:
            t = await cur.fetchone()
        new_status = 0 if t["is_active"] else 1
        await dbc.execute("UPDATE tariffs SET is_active=? WHERE id=?", (new_status, tariff_id))
        await dbc.commit()
    label = "✅ Включён" if new_status else "❌ Отключён"
    await call.answer(label)
    await show_tariff_card(call.message, tariff_id, edit=True)


@router.callback_query(F.data.startswith("tariff_edit_menu"))
async def cb_tariff_edit_menu(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    tariff_id = int(call.data.split(":")[1])
    await call.message.edit_text("<b>Что изменить?</b>", reply_markup=tariff_edit_fields_kb(tariff_id), parse_mode="HTML")


FIELD_LABELS = {
    "name": "название",
    "description": "описание",
    "days": "количество дней",
    "price": "цену",
}


@router.callback_query(F.data.startswith("tariff_edit_field"))
async def cb_tariff_edit_field(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    _, tariff_id, field = call.data.split(":")
    await state.update_data(edit_tariff_id=int(tariff_id), edit_field=field)
    await state.set_state(TariffEditState.value)
    await call.message.edit_text(
        f"Введи новое {FIELD_LABELS.get(field, field)}:",
        reply_markup=back_kb(f"tariff_edit_menu:{tariff_id}")
    )


@router.message(TariffEditState.value)
async def handle_tariff_field_value(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    tariff_id = data["edit_tariff_id"]
    field = data["edit_field"]
    value = message.text.strip()
    await state.set_state(None)
    if field in ("days", "price"):
        try:
            value = int(value) if field == "days" else float(value)
        except ValueError:
            await message.answer(
                "❌ Введи корректное число.",
                reply_markup=back_kb(f"tariff_edit_menu:{tariff_id}")
            )
            await state.set_state(TariffEditState.value)
            return
    await db.update_tariff(tariff_id, **{field: value})
    await message.answer("✅ Сохранено!")
    await show_tariff_card(message, tariff_id, edit=False)


def currency_choice_kb(callback_prefix: str):
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="₽ RUB", callback_data=f"{callback_prefix}:RUB"),
        InlineKeyboardButton(text="€ EUR", callback_data=f"{callback_prefix}:EUR"),
        InlineKeyboardButton(text="$ USD", callback_data=f"{callback_prefix}:USD"),
    )
    return builder.as_markup()


@router.callback_query(F.data.startswith("tariff_edit_currency"))
async def cb_tariff_edit_currency(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    tariff_id = int(call.data.split(":")[1])
    await call.message.edit_text(
        "Выбери валюту:",
        reply_markup=currency_choice_kb(f"tariff_set_currency:{tariff_id}")
    )


@router.callback_query(F.data.startswith("tariff_set_currency"))
async def cb_tariff_set_currency(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    parts = call.data.split(":")
    tariff_id, currency = int(parts[1]), parts[2]
    await db.update_tariff(tariff_id, currency=currency)
    symbol = CURRENCY_SYMBOLS.get(currency, currency)
    await call.answer(f"✅ Валюта: {symbol} {currency}")
    await show_tariff_card(call.message, tariff_id, edit=True)


@router.callback_query(F.data.startswith("tariff_delete_confirm"))
async def cb_tariff_delete_confirm(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    tariff_id = int(call.data.split(":")[1])
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"tariff_delete:{tariff_id}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data=f"admin_edit_tariff:{tariff_id}")
    )
    await call.message.edit_text(
        "<b>Удалить тариф?</b>\nВсе привязанные подписки останутся, но новые нельзя будет купить.",
        reply_markup=builder.as_markup(), parse_mode="HTML"
    )


@router.callback_query(F.data.startswith("tariff_delete:"))
async def cb_tariff_delete(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    tariff_id = int(call.data.split(":")[1])
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        await dbc.execute("DELETE FROM tariffs WHERE id=?", (tariff_id,))
        await dbc.commit()
    await call.answer("🗑 Удалён")
    await cb_admin_tariffs(call)


@router.callback_query(F.data == "admin_add_tariff")
async def cb_admin_add_tariff(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    await state.set_state(TariffAddState.name)
    await call.message.edit_text(
        "<b>1/4 · Название тарифа</b>\n\n<i>Пример: 1 месяц</i>",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )


@router.message(TariffAddState.name)
async def tariff_add_name(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.update_data(new_name=message.text.strip())
    await state.set_state(TariffAddState.description)
    await message.answer("2/4 · <b>Описание</b> (можно пропустить, отправь «-»)\n\n<i>Пример: Доступ на 30 дней</i>", parse_mode="HTML")


@router.message(TariffAddState.description)
async def tariff_add_description(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    desc = message.text.strip()
    await state.update_data(new_description=None if desc == "-" else desc)
    await state.set_state(TariffAddState.days)
    await message.answer("3/4 · <b>Количество дней:</b>", parse_mode="HTML")


@router.message(TariffAddState.days)
async def tariff_add_days(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        days = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введи число.")
        return
    await state.update_data(new_days=days)
    await state.set_state(TariffAddState.price)
    await message.answer("4/4 · <b>Цена</b> (введи 0 для бесплатного):", parse_mode="HTML")


@router.message(TariffAddState.price)
async def tariff_add_price(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        price = float(message.text.strip())
    except ValueError:
        await message.answer("❌ Введи число.")
        return
    await state.update_data(new_price=price)
    await state.set_state(TariffAddState.currency)
    await message.answer("Выбери валюту:", reply_markup=currency_choice_kb("tariff_add_currency"))


@router.callback_query(F.data.startswith("tariff_add_currency"), TariffAddState.currency)
async def tariff_add_currency(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    currency = call.data.split(":")[1]
    data = await state.get_data()
    await state.set_state(None)
    tariff_id = await db.add_tariff(
        name=data["new_name"],
        description=data["new_description"],
        days=data["new_days"],
        price=data["new_price"]
    )
    await db.update_tariff(tariff_id, currency=currency)
    await call.message.edit_text("✅ Тариф добавлен!")
    await show_tariff_card(call.message, tariff_id, edit=False)


# ─── PAYMENT METHODS ─────────────────────────────────────────────────────────

class PaymentMethodState(StatesGroup):
    name = State()
    currency = State()
    details = State()
    is_link = State()


class PmEditState(StatesGroup):
    details = State()
    name = State()


def payment_methods_admin_kb(methods: list):
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    for m in methods:
        status = "✅" if m["is_active"] else "❌"
        link_icon = " 🔗" if m["is_link"] else ""
        currency = m["currency"] or "RUB"
        symbol = CURRENCY_SYMBOLS.get(currency, currency)
        builder.row(InlineKeyboardButton(
            text=f"{status}{link_icon} {m['name']} · {symbol}",
            callback_data=f"pm_card:{m['id']}"
        ))
    builder.row(InlineKeyboardButton(text="➕ Добавить метод", callback_data="pm_add"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="admin_settings"))
    return builder.as_markup()


def pm_card_kb(method_id: int, is_active: int):
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    toggle_text = "❌ Отключить" if is_active else "✅ Включить"
    builder.row(InlineKeyboardButton(text=toggle_text, callback_data=f"pm_toggle:{method_id}"))
    builder.row(InlineKeyboardButton(text="✏️ Изменить реквизиты", callback_data=f"pm_edit_details:{method_id}"))
    builder.row(InlineKeyboardButton(text="📝 Изменить название", callback_data=f"pm_edit_name:{method_id}"))
    builder.row(InlineKeyboardButton(text="💱 Изменить валюту", callback_data=f"pm_edit_currency:{method_id}"))
    builder.row(InlineKeyboardButton(text="🗑 Удалить", callback_data=f"pm_delete_confirm:{method_id}"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="admin_payment_methods"))
    return builder.as_markup()


@router.callback_query(F.data == "admin_payment_methods")
async def cb_payment_methods(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    methods = await db.get_all_payment_methods()
    text = "<b>💳 Методы оплаты</b>\n\n"
    if methods:
        for m in methods:
            status = "✅" if m["is_active"] else "❌"
            currency = m["currency"] or "RUB"
            symbol = CURRENCY_SYMBOLS.get(currency, currency)
            preview = m["details"][:60] + "..." if len(m["details"]) > 60 else m["details"]
            text += f"{status} <b>{m['name']}</b> · {symbol}\n<code>{preview}</code>\n\n"
    else:
        text += "Нет методов оплаты!"
    await call.message.edit_text(text, reply_markup=payment_methods_admin_kb(methods), parse_mode="HTML")


@router.callback_query(F.data.startswith("pm_card"))
async def cb_pm_card(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    method_id = int(call.data.split(":")[1])
    await show_pm_card(call.message, method_id, edit=True)


async def show_pm_card(target, method_id: int, edit: bool = True):
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        dbc.row_factory = aiosqlite.Row
        async with dbc.execute("SELECT * FROM payment_methods WHERE id=?", (method_id,)) as cur:
            m = await cur.fetchone()
    if not m:
        return
    status = "✅ Активен" if m["is_active"] else "❌ Отключён"
    link_type = "🔗 Ссылка" if m["is_link"] else "💳 Реквизиты"
    currency = m["currency"] or "RUB"
    symbol = CURRENCY_SYMBOLS.get(currency, currency)
    text = (
        f"💳 <b>{m['name']}</b>\n"
        f"💱 Валюта: <b>{symbol} {currency}</b>\n"
        f"🏷 Тип: {link_type}\n"
        f"📊 {status}\n\n"
        f"<code>{m['details']}</code>"
    )
    kb = pm_card_kb(method_id, m["is_active"])
    if edit:
        await target.edit_text(text, reply_markup=kb, parse_mode="HTML")
    else:
        await target.answer(text, reply_markup=kb, parse_mode="HTML")


@router.callback_query(F.data.startswith("pm_toggle"))
async def cb_pm_toggle(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    method_id = int(call.data.split(":")[1])
    await db.toggle_payment_method(method_id)
    await show_pm_card(call.message, method_id, edit=True)


@router.callback_query(F.data.startswith("pm_delete_confirm"))
async def cb_pm_delete_confirm(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    method_id = int(call.data.split(":")[1])
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="✅ Да, удалить", callback_data=f"pm_delete:{method_id}"),
        InlineKeyboardButton(text="❌ Отмена", callback_data=f"pm_card:{method_id}")
    )
    await call.message.edit_text("<b>Удалить метод оплаты?</b>", reply_markup=builder.as_markup(), parse_mode="HTML")


@router.callback_query(F.data.startswith("pm_delete:"))
async def cb_pm_delete(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    method_id = int(call.data.split(":")[1])
    await db.delete_payment_method(method_id)
    await call.answer("🗑 Удалён")
    await cb_payment_methods(call)


@router.callback_query(F.data.startswith("pm_edit_details"))
async def cb_pm_edit_details(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    method_id = int(call.data.split(":")[1])
    await state.update_data(pm_edit_id=method_id)
    await state.set_state(PmEditState.details)
    await call.message.edit_text(
        "Введи новые реквизиты:\n<i>Для ссылки на оплату — начни с https://</i>",
        reply_markup=back_kb(f"pm_card:{method_id}"), parse_mode="HTML"
    )


@router.message(PmEditState.details)
async def handle_pm_edit_details(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    method_id = data["pm_edit_id"]
    new_details = message.text.strip()
    is_link = 1 if new_details.startswith("http") else 0
    await state.set_state(None)
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        await dbc.execute(
            "UPDATE payment_methods SET details=?, is_link=? WHERE id=?",
            (new_details, is_link, method_id)
        )
        await dbc.commit()
    await message.answer("✅ Реквизиты обновлены!")
    await show_pm_card(message, method_id, edit=False)


@router.callback_query(F.data.startswith("pm_edit_name"))
async def cb_pm_edit_name(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    method_id = int(call.data.split(":")[1])
    await state.update_data(pm_edit_id=method_id)
    await state.set_state(PmEditState.name)
    await call.message.edit_text("Введи новое название:", reply_markup=back_kb(f"pm_card:{method_id}"))


@router.message(PmEditState.name)
async def handle_pm_edit_name(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    method_id = data["pm_edit_id"]
    await state.set_state(None)
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        await dbc.execute("UPDATE payment_methods SET name=? WHERE id=?", (message.text.strip(), method_id))
        await dbc.commit()
    await message.answer("✅ Название обновлено!")
    await show_pm_card(message, method_id, edit=False)


@router.callback_query(F.data.startswith("pm_edit_currency"))
async def cb_pm_edit_currency(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    method_id = int(call.data.split(":")[1])
    await call.message.edit_text(
        "Выбери валюту:",
        reply_markup=currency_choice_kb(f"pm_set_currency:{method_id}")
    )


@router.callback_query(F.data.startswith("pm_set_currency"))
async def cb_pm_set_currency(call: CallbackQuery):
    if not is_admin(call.from_user.id):
        return
    parts = call.data.split(":")
    method_id, currency = int(parts[1]), parts[2]
    import aiosqlite as aiosqlite
    async with aiosqlite.connect(config.DB_PATH) as dbc:
        await dbc.execute("UPDATE payment_methods SET currency=? WHERE id=?", (currency, method_id))
        await dbc.commit()
    symbol = CURRENCY_SYMBOLS.get(currency, currency)
    await call.answer(f"✅ Валюта: {symbol} {currency}")
    await show_pm_card(call.message, method_id, edit=True)


@router.callback_query(F.data == "pm_add")
async def cb_pm_add(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    await state.set_state(PaymentMethodState.name)
    await call.message.edit_text(
        "<b>1/4 · Название метода оплаты</b>\n\n<i>Пример: Сбербанк, IBAN EUR, Крипто USDT</i>",
        reply_markup=cancel_kb(), parse_mode="HTML"
    )


@router.message(PaymentMethodState.name)
async def pm_set_name(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.update_data(pm_name=message.text.strip())
    await state.set_state(PaymentMethodState.currency)
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    for cur in ["RUB", "EUR", "USD"]:
        symbol = CURRENCY_SYMBOLS[cur]
        builder.row(InlineKeyboardButton(text=f"{symbol} {cur}", callback_data=f"pm_currency:{cur}"))
    await message.answer("2/4 · <b>Валюта:</b>", reply_markup=builder.as_markup(), parse_mode="HTML")


@router.callback_query(F.data.startswith("pm_currency"), PaymentMethodState.currency)
async def pm_set_currency(call: CallbackQuery, state: FSMContext):
    currency = call.data.split(":")[1]
    await state.update_data(pm_currency=currency)
    await state.set_state(PaymentMethodState.details)
    await call.message.edit_text(
        "3/4 · <b>Реквизиты</b>\n\n"
        "<i>Номер карты, IBAN, ссылка на оплату и т.д.</i>\n\n"
        f"Пример: <code>https://www.tinkoff.ru/rm/ivanov.xxx</code>\n"
        f"Или: <code>IBAN DE89 3704... COBADEFF</code>",
        parse_mode="HTML"
    )


@router.message(PaymentMethodState.details)
async def pm_set_details(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    details = message.text.strip()
    is_link = 1 if details.startswith("http") else 0
    await state.update_data(pm_details=details, pm_is_link=is_link)
    await state.set_state(PaymentMethodState.is_link)
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="🔗 Ссылка на оплату", callback_data="pm_type:1"),
        InlineKeyboardButton(text="💳 Реквизиты", callback_data="pm_type:0")
    )
    link_detected = " (определено автоматически)" if is_link else ""
    await message.answer(f"4/4 · Тип реквизитов?{link_detected}", reply_markup=builder.as_markup())


@router.callback_query(F.data.startswith("pm_type"), PaymentMethodState.is_link)
async def pm_confirm(call: CallbackQuery, state: FSMContext):
    if not is_admin(call.from_user.id):
        return
    is_link = int(call.data.split(":")[1])
    data = await state.get_data()
    await state.set_state(None)
    method_id = await db.add_payment_method(
        name=data["pm_name"],
        currency=data["pm_currency"],
        details=data["pm_details"],
        is_link=is_link
    )
    type_label = "🔗 ссылка" if is_link else "💳 реквизиты"
    currency = data["pm_currency"]
    symbol = CURRENCY_SYMBOLS.get(currency, currency)
    await call.message.edit_text(
        f"✅ <b>Метод добавлен!</b>\n"
        f"📝 {data['pm_name']}\n"
        f"💱 {symbol} {currency}\n"
        f"🏷 {type_label}",
        reply_markup=back_kb("admin_payment_methods"), parse_mode="HTML"
    )
